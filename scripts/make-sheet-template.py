#!/usr/bin/env python3
"""Generate the Polish employee data-collection workbook for Shiftwise.

The workbook is uploaded to Google Drive, filled in by the employees, downloaded
again as .xlsx (or .csv) and imported through Workers -> Import from sheet.

The tab names, header labels and cell codes written here are the ones the
importer understands. They are mirrored in src/sheet-io.ts (HEADER_SYNONYMS,
PERIOD_VALUES, AVAILABILITY_CODES); change both sides together. Anything the
importer ignores -- titles, legends, the Razem column -- is free to move.

    python3 scripts/make-sheet-template.py --month 2026-08 --rows 30
    python3 scripts/make-sheet-template.py --month 2026-08 --sample

scripts/grafik-arkusz.gs builds the same layout natively inside Google Sheets.
"""

import argparse
import calendar
import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# The app's own palette, so the sheet and the scheduler look like one product.
INK = '10281E'
ACCENT = '1D7645'
MINT = 'EDF9F1'
LINE = 'DCE6E1'
MUTED = '6B7A72'
BAND = 'F8FBF9'
WEEKEND = 'EFF3F1'
PAPER = 'FFFFFF'

# kod: (tło, kolor tekstu, krótka etykieta do legendy, pełny opis do instrukcji)
CODE_STYLES = {
    'X': ('FCE3E3', '8C2F2F', 'cały dzień', 'niedostępny przez cały dzień (urlop, L4)'),
    'D': ('FDEFD5', '8A5A12', 'rano / dzień', 'niedostępny rano i w dzień — noce są w porządku'),
    'N': ('E2E4FA', '3A3E8F', 'noc', 'niedostępny w nocy — dni są w porządku'),
}

WEEKDAYS_PL = ['Pn', 'Wt', 'Śr', 'Cz', 'Pt', 'So', 'Nd']
MONTHS_PL = ['styczeń', 'luty', 'marzec', 'kwiecień', 'maj', 'czerwiec',
             'lipiec', 'sierpień', 'wrzesień', 'październik', 'listopad', 'grudzień']

PERIOD_CHOICES = ['Dzień', 'Noc', 'Bez preferencji']
PAIR_CHOICES = ['Nie', 'Dowolny', '08:00 → 08:00', '20:00 → 20:00']
YES_NO = ['Tak', 'Nie']

WORKER_COLUMNS = [
    ('Imię i nazwisko', 28, 'Tak samo jak w zakładce Dostępność — najlepiej wybierz siebie z listy.'),
    ('Godziny docelowe', 15, 'Ile godzin chcesz przepracować w tym miesiącu. Liczba, np. 160.'),
    ('Preferowana pora', 17, 'Dzień, Noc albo Bez preferencji.'),
    ('Dyżur 24h', 17, 'Czy bierzesz dyżury 24-godzinne i w jakim układzie.'),
]

# Zakładka tylko dla osoby układającej grafik — arkusz jest chroniony.
ADMIN_COLUMNS = [
    ('Imię i nazwisko', 28, 'Musi brzmieć tak samo jak w zakładce Pracownicy.'),
    ('Kategorie', 24, 'Kwalifikacje, po przecinku. Puste = General.'),
    ('Uprawnienia kierownika', 20, 'Tak, jeśli ta osoba może pełnić zmianę kierownika.'),
    ('Kierownik domyślny', 18, 'Tak tylko dla jednej osoby w zespole.'),
    ('Uwagi', 34, 'Notatki osoby układającej grafik — program je pomija.'),
]

SAMPLE_WORKERS = [
    ('Łukasz Zieliński', 160, 'Dzień', 'Nie'),
    ('Agnieszka Wójcik', 152, 'Noc', 'Dowolny'),
    ('Paweł Nowak', 168, 'Bez preferencji', '08:00 → 08:00'),
    ('Maja Dąbrowska', 144, 'Dzień', 'Nie'),
]
SAMPLE_ADMIN = [
    ('Łukasz Zieliński', 'General', 'Tak', 'Tak', ''),
    ('Agnieszka Wójcik', 'General', 'Nie', 'Nie', 'Wolę bloki nocne.'),
    ('Paweł Nowak', 'General, Nursing', 'Tak', 'Nie', ''),
    ('Maja Dąbrowska', 'General', 'Nie', 'Nie', 'Urlop w drugim tygodniu.'),
]
# {imię: {dzień: kod}} — pokazuje pracownikom, jak wygląda wypełniony arkusz.
SAMPLE_AVAILABILITY = {
    'Łukasz Zieliński': {3: 'N', 4: 'N', 17: 'X'},
    'Agnieszka Wójcik': {8: 'D', 9: 'D', 22: 'X', 23: 'X'},
    'Paweł Nowak': {12: 'X'},
    'Maja Dąbrowska': {10: 'X', 11: 'X', 12: 'X', 13: 'X', 14: 'X'},
}

THIN = Side(style='thin', color=LINE)


def title_font(size=18):
    return Font(bold=True, size=size, color=INK)


def muted_font(size=11, italic=False):
    return Font(size=size, color=MUTED, italic=italic)


def fill(colour):
    return PatternFill('solid', fgColor=colour)


def paint_header(cell):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = fill(INK)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='medium', color=ACCENT))


def month_name(year, month):
    return '%s %d' % (MONTHS_PL[month - 1], year)


def build_instructions(sheet, year, month):
    sheet.sheet_properties.tabColor = ACCENT
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 104

    def write(row, code, text, style):
        marker = sheet.cell(row=row, column=1, value=code)
        body = sheet.cell(row=row, column=2, value=text)
        if style == 'title':
            body.font = title_font(20)
            sheet.row_dimensions[row].height = 30
        elif style == 'lead':
            body.font = muted_font(12)
            sheet.row_dimensions[row].height = 20
        elif style == 'step':
            marker.font = Font(bold=True, size=12, color='FFFFFF')
            marker.fill = fill(ACCENT)
            marker.alignment = Alignment(horizontal='center', vertical='center')
            body.font = Font(bold=True, size=12, color=INK)
            sheet.row_dimensions[row].height = 22
        elif style == 'body':
            body.font = Font(size=11, color=INK)
            body.alignment = Alignment(wrap_text=True, vertical='top')
            sheet.row_dimensions[row].height = 15 * (1 + len(text) // 100)
        elif style == 'code':
            background, ink = CODE_STYLES.get(code, (MINT, INK, '', ''))[:2]
            marker.font = Font(bold=True, size=12, color=ink)
            marker.fill = fill(background)
            marker.alignment = Alignment(horizontal='center', vertical='center')
            marker.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            body.font = Font(size=11, color=INK)
            sheet.row_dimensions[row].height = 20
        elif style == 'head':
            body.font = Font(bold=True, size=12, color=ACCENT)
            sheet.row_dimensions[row].height = 26
        elif style == 'note':
            body.font = muted_font(10, italic=True)
            body.alignment = Alignment(wrap_text=True, vertical='top')
            sheet.row_dimensions[row].height = 15 * (1 + len(text) // 100)
        else:
            sheet.row_dimensions[row].height = 8

    rows = [
        ('', 'Grafik — dane od pracowników', 'title'),
        ('', '%s · wypełnij swój wiersz w obu zakładkach, zajmuje to około dwóch minut' % month_name(year, month), 'lead'),
        ('', '', 'gap'),
        ('1', 'Zakładka „Pracownicy”', 'step'),
        ('', 'Znajdź swoje imię i nazwisko albo dopisz je w pierwszym wolnym wierszu. Uzupełnij, ile godzin '
             'chcesz przepracować, jaką porę dnia wolisz i czy bierzesz dyżury 24h. Komórki z listą rozwijaną '
             'wypełnij wyłącznie wartościami z listy.', 'body'),
        ('', '', 'gap'),
        ('2', 'Zakładka „Dostępność”', 'step'),
        ('', 'W swoim wierszu zaznacz tylko te dni, w których NIE możesz pracować. Puste pole znaczy, '
             'że jesteś dostępny — przy dniach, w które możesz pracować, nie wpisuj niczego.', 'body'),
        ('', '', 'gap'),
        ('', 'Zakładka „Administrator”', 'step'),
        ('', 'Należy do osoby układającej grafik: kwalifikacje, uprawnienia kierownika i notatki. '
             'Arkusz jest chroniony — pracownicy go nie zmieniają.', 'body'),
        ('', '', 'gap'),
        ('', 'Kody dostępności', 'head'),
    ]
    for code, (_, _, _, description) in CODE_STYLES.items():
        rows.append((code, description, 'code'))
    rows += [
        ('', 'puste pole — jestem dostępny', 'body'),
        ('', '', 'gap'),
        ('', 'O czym pamiętać', 'head'),
        ('', 'Imię i nazwisko musi brzmieć tak samo w obu zakładkach — w „Dostępności” wybierz je z listy.', 'body'),
        ('', 'Nie zmieniaj nazw zakładek ani nagłówków kolumn — po nich program rozpoznaje dane.', 'body'),
        ('', 'Możesz dopisywać wiersze na dole i dodawać własne kolumny; program pomija to, czego nie zna.', 'body'),
        ('', '', 'gap'),
        ('', 'Dla osoby układającej grafik', 'head'),
        ('', 'Gdy zespół skończy: Plik → Pobierz → Microsoft Excel (.xlsx), a następnie w programie Shiftwise '
             'zakładka Workers → Import from sheet. Przed zapisem zobaczysz podgląd wszystkich zmian.', 'body'),
        ('', 'Miesiąc, którego dotyczy arkusz, jest w komórce B1 zakładki „Dostępność”. Aby użyć tego samego '
             'arkusza w kolejnym miesiącu, zmień tę komórkę i wyczyść siatkę.', 'note'),
    ]
    for index, (code, text, style) in enumerate(rows, start=1):
        write(index, code, text, style)


def build_workers(sheet, rows, year, month, sample=False):
    sheet.sheet_properties.tabColor = INK
    sheet.sheet_view.showGridLines = False

    sheet.cell(row=1, column=1, value='Pracownicy').font = title_font(15)
    subtitle = sheet.cell(row=2, column=1, value='Jeden wiersz na osobę · %s · wypełniają pracownicy' % month_name(year, month))
    subtitle.font = muted_font(11)

    header_row, first = 3, 4
    last = first + rows - 1
    for index, (label, width, hint) in enumerate(WORKER_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
        header = sheet.cell(row=header_row, column=index, value=label)
        paint_header(header)
        # Hints live in comments, not in a row of their own: a text row under the
        # header would be read back as a worker called "Ile godzin chcesz…".
        header.comment = Comment(hint, 'Grafik', height=80, width=250)
    sheet.row_dimensions[header_row].height = 32

    for row in range(first, last + 1):
        banded = (row - first) % 2 == 1
        for column in range(1, len(WORKER_COLUMNS) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = Border(bottom=THIN, right=THIN)
            cell.alignment = Alignment(vertical='center', horizontal='left' if column == 1 else 'center')
            cell.fill = fill(BAND if banded else PAPER)
            if column == 2:
                cell.number_format = '0 "h"'
        # Kategorie stays empty on purpose: the importer falls back to General.
        sheet.row_dimensions[row].height = 18

    validations = [
        ('C', PERIOD_CHOICES, 'Preferowana pora', 'Wybierz: Dzień, Noc albo Bez preferencji.'),
        ('D', PAIR_CHOICES, 'Dyżur 24h', 'Nie, Dowolny albo konkretny układ 24-godzinny.'),
    ]
    for letter, choices, caption, prompt in validations:
        rule = DataValidation(type='list', formula1='"' + ','.join(choices) + '"', allow_blank=True,
                              showDropDown=False, errorTitle=caption, error=prompt,
                              promptTitle=caption, prompt=prompt)
        sheet.add_data_validation(rule)
        rule.add('%s%d:%s%d' % (letter, first, letter, last))

    hours = DataValidation(type='whole', operator='between', formula1=0, formula2=400, allow_blank=True,
                           errorTitle='Godziny docelowe', error='Podaj liczbę godzin z zakresu 0–400.',
                           promptTitle='Godziny docelowe', prompt='Ile godzin chcesz przepracować w tym miesiącu.')
    sheet.add_data_validation(hours)
    hours.add('B%d:B%d' % (first, last))

    if sample:
        for offset, values in enumerate(SAMPLE_WORKERS):
            for column, value in enumerate(values, start=1):
                if value != '':
                    sheet.cell(row=first + offset, column=column).value = value

    sheet.freeze_panes = 'A%d' % first
    return first, last


def build_administrator(sheet, rows, year, month, sample=False):
    """Kwalifikacje i role kierownika — tylko dla osoby układającej grafik."""
    sheet.sheet_properties.tabColor = MUTED
    sheet.sheet_view.showGridLines = False

    sheet.cell(row=1, column=1, value='Administrator').font = title_font(15)
    subtitle = sheet.cell(row=2, column=1,
                          value='Kwalifikacje i role kierownika · %s · wypełnia osoba układająca grafik' % month_name(year, month))
    subtitle.font = muted_font(11)

    header_row, first = 3, 4
    last = first + rows - 1
    for index, (label, width, hint) in enumerate(ADMIN_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
        header = sheet.cell(row=header_row, column=index, value=label)
        paint_header(header)
        header.comment = Comment(hint, 'Grafik', height=80, width=250)
    sheet.row_dimensions[header_row].height = 32

    for row in range(first, last + 1):
        banded = (row - first) % 2 == 1
        for column in range(1, len(ADMIN_COLUMNS) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = Border(bottom=THIN, right=THIN)
            cell.alignment = Alignment(vertical='center', horizontal='left' if column in (1, 2, 5) else 'center')
            cell.fill = fill(BAND if banded else PAPER)
        sheet.row_dimensions[row].height = 18

    for letter in ('C', 'D'):
        rule = DataValidation(type='list', formula1='"' + ','.join(YES_NO) + '"', allow_blank=True,
                              showDropDown=False, errorTitle='Tak albo Nie', error='Wpisz Tak albo Nie.')
        sheet.add_data_validation(rule)
        rule.add('%s%d:%s%d' % (letter, first, letter, last))

    names = DataValidation(type='list', formula1='=Pracownicy!$A$4:$A$%d' % last, allow_blank=True,
                           showDropDown=False, promptTitle='Imię i nazwisko',
                           prompt='Wybierz osobę z zakładki Pracownicy.')
    sheet.add_data_validation(names)
    names.add('A%d:A%d' % (first, last))

    if sample:
        for offset, values in enumerate(SAMPLE_ADMIN):
            for column, value in enumerate(values, start=1):
                if value != '':
                    sheet.cell(row=first + offset, column=column).value = value

    # Guardrail, nie zamek: w Excelu ochronę bez hasła zdejmuje się jednym kliknięciem,
    # a prawdziwą blokadę daje dopiero scripts/grafik-arkusz.gs w Arkuszach Google.
    sheet.protection.sheet = True
    sheet.protection.enable()
    sheet.freeze_panes = 'A%d' % first


def build_availability(sheet, year, month, rows, worker_rows, sample=False):
    sheet.sheet_properties.tabColor = ACCENT
    sheet.sheet_view.showGridLines = False
    days = calendar.monthrange(year, month)[1]
    header_row, first = 4, 5
    last = first + rows - 1
    last_day_column = days + 1
    total_column = days + 2

    label = sheet.cell(row=1, column=1, value='Miesiąc')
    label.font = Font(bold=True, size=11, color=MUTED)
    label.alignment = Alignment(horizontal='right', vertical='center')
    value = sheet.cell(row=1, column=2, value='%04d-%02d' % (year, month))
    value.font = Font(bold=True, size=12, color=ACCENT)
    value.fill = fill(MINT)
    value.alignment = Alignment(horizontal='center', vertical='center')
    value.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    sheet.row_dimensions[1].height = 22

    banner = sheet.cell(row=1, column=4, value='Zaznacz tylko te dni, w których NIE możesz pracować — %s' % month_name(year, month))
    banner.font = Font(bold=True, size=12, color=INK)
    sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=min(last_day_column, 20))

    # Legend chips: the code cell is painted exactly like a filled-in cell.
    column = 4
    for code, (background, ink, caption_text, _) in CODE_STYLES.items():
        chip = sheet.cell(row=2, column=column, value=code)
        chip.font = Font(bold=True, size=11, color=ink)
        chip.fill = fill(background)
        chip.alignment = Alignment(horizontal='center', vertical='center')
        chip.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        caption = sheet.cell(row=2, column=column + 1, value=caption_text)
        caption.font = muted_font(10)
        column += 6
    trailing = sheet.cell(row=2, column=column, value='puste = dostępny')
    trailing.font = muted_font(10)
    sheet.row_dimensions[2].height = 18

    weekday_label = sheet.cell(row=3, column=1, value='Dzień tygodnia')
    weekday_label.font = muted_font(9, italic=True)
    weekday_label.alignment = Alignment(horizontal='right', vertical='center')

    name_header = sheet.cell(row=header_row, column=1, value='Imię i nazwisko')
    paint_header(name_header)
    name_header.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    sheet.column_dimensions['A'].width = 28
    sheet.row_dimensions[header_row].height = 22

    weekend_columns = []
    for day in range(1, days + 1):
        index = day + 1
        sheet.column_dimensions[get_column_letter(index)].width = 4.2
        weekday = dt.date(year, month, day).weekday()
        top = sheet.cell(row=3, column=index, value=WEEKDAYS_PL[weekday])
        top.font = Font(size=9, color=MUTED, bold=weekday >= 5)
        top.alignment = Alignment(horizontal='center')
        if weekday >= 5:
            weekend_columns.append(index)
            top.fill = fill(WEEKEND)
        cell = sheet.cell(row=header_row, column=index, value=day)
        paint_header(cell)

    total_header = sheet.cell(row=header_row, column=total_column, value='Razem')
    paint_header(total_header)
    sheet.column_dimensions[get_column_letter(total_column)].width = 8

    for row in range(first, last + 1):
        banded = (row - first) % 2 == 1
        name_cell = sheet.cell(row=row, column=1)
        name_cell.border = Border(bottom=THIN, right=THIN)
        name_cell.alignment = Alignment(vertical='center', indent=1)
        name_cell.fill = fill(BAND if banded else PAPER)
        for index in range(2, days + 2):
            cell = sheet.cell(row=row, column=index)
            cell.border = Border(bottom=THIN, right=THIN)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill(WEEKEND if index in weekend_columns else (BAND if banded else PAPER))
        total = sheet.cell(row=row, column=total_column,
                           value='=COUNTA(B%d:%s%d)' % (row, get_column_letter(last_day_column), row))
        total.font = Font(bold=True, size=10, color=MUTED)
        total.alignment = Alignment(horizontal='center', vertical='center')
        total.border = Border(bottom=THIN, right=THIN)
        sheet.row_dimensions[row].height = 18

    names = DataValidation(type='list', formula1='=Pracownicy!$A$%d:$A$%d' % worker_rows, allow_blank=True,
                           showDropDown=False, promptTitle='Imię i nazwisko',
                           prompt='Wybierz siebie z listy z zakładki Pracownicy.')
    sheet.add_data_validation(names)
    names.add('A%d:A%d' % (first, last))

    codes = DataValidation(type='list', formula1='"X,D,N"', allow_blank=True, showDropDown=False,
                           errorTitle='Kod dostępności',
                           error='Dozwolone kody: X (cały dzień), D (rano / dzień), N (noc). Puste pole = jestem dostępny.',
                           promptTitle='Niedostępność', prompt='X = cały dzień, D = rano / dzień, N = noc.')
    sheet.add_data_validation(codes)
    grid = 'B%d:%s%d' % (first, get_column_letter(last_day_column), last)
    codes.add(grid)

    for code, (background, ink, _, _) in CODE_STYLES.items():
        sheet.conditional_formatting.add(grid, CellIsRule(
            operator='equal', formula=['"%s"' % code],
            fill=fill(background), font=Font(bold=True, color=ink)))

    if sample:
        for offset, (name, marks) in enumerate(SAMPLE_AVAILABILITY.items()):
            sheet.cell(row=first + offset, column=1).value = name
            for day, code in marks.items():
                if day <= days:
                    sheet.cell(row=first + offset, column=day + 1).value = code

    sheet.freeze_panes = 'B%d' % first
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def build_workbook(year, month, rows, sample=False):
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = 'Instrukcja'
    build_instructions(instructions, year, month)
    workers = workbook.create_sheet('Pracownicy')
    worker_rows = build_workers(workers, rows, year, month, sample)
    availability = workbook.create_sheet('Dostępność')
    build_availability(availability, year, month, rows, worker_rows, sample)
    administrator = workbook.create_sheet('Administrator')
    build_administrator(administrator, rows, year, month, sample)
    return workbook


def next_month(today):
    return (today.year + 1, 1) if today.month == 12 else (today.year, today.month + 1)


def main():
    default_year, default_month = next_month(dt.date.today())
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--month', default='%04d-%02d' % (default_year, default_month),
                        help='Miesiąc arkusza w formacie RRRR-MM (domyślnie następny miesiąc).')
    parser.add_argument('--rows', type=int, default=30, help='Liczba wierszy pracowników.')
    parser.add_argument('--out', default=None, help='Ścieżka pliku wyjściowego .xlsx.')
    parser.add_argument('--sample', action='store_true', help='Wypełnij arkusz przykładowymi danymi.')
    options = parser.parse_args()

    year, month = (int(part) for part in options.month.split('-'))
    suffix = '-przyklad' if options.sample else ''
    output = options.out or os.path.join('templates', 'Grafik-dane-%04d-%02d%s.xlsx' % (year, month, suffix))
    os.makedirs(os.path.dirname(output) or '.', exist_ok=True)
    build_workbook(year, month, options.rows, options.sample).save(output)
    print('Zapisano %s' % output)


if __name__ == '__main__':
    main()
